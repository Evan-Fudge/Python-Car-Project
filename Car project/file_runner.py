from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import styles
from openpyxl import formatting


wb = load_workbook("master_car_mileage_report.xlsx")
ws = wb['Sheet1']
mwb = load_workbook("MANUFACTURER_MILEAGEREX.xlsx")
ws3 = mwb['Sheet1']
#
#
#
# for row in range(1,68):
#     for col in range(1,8):
#         char = get_column_letter(col)


# for row in range(2,68):
#     for col in range(5,6):
#         char = get_column_letter(col)
#         charv = get_column_letter(col-2)
#         ws[char + str(row)] = f"=Vlookup({charv + str(row)},Sheet2!A:B,2,False)"""
#
# for row in range(2,68):
#   for col in range(5,6):
#       char = get_column_letter(col)
#       charv = get_column_letter(col - 2)
#       ws[char + str(row)] = f"=Xlookup({charv + str(row)},{ws2['A:A']},{ws2['B:B']},"")"
#
#
# for row in range(2,68):
#     for col in range(6,7):
#         char = get_column_letter(col)
#         max = get_column_letter(col-1)
#         subtractor = get_column_letter(col-2)
#         ws[char + str(row)] = f"={max+str(row)}-{subtractor+str(row)}"



#
# for row in ws3.iter_rows():
#     make_name = row[0].value
#     mileage = row[1].value
#     make_row = row[0].row
#     for j in ws.iter_rows():
#         if j[2].value == make_name:
#             ws.cell(row=30, column= 5).value = mileage


for row in ws.iter_rows():
    make = row[2].value
    make_row = row[0].row
    for j in ws3.iter_rows():
        if j[0].value == make:
            ws.cell(row= make_row, column=5).value = j[1].value


for row in ws.iter_rows(min_row=2):
    miles = row[3].value
    max_miles = row[4].value
    miles_left = row[5].row
    for j in ws.iter_rows():
        if max_miles != None:
            ws.cell(row=miles_left,column=6).value = max_miles-miles

ws.insert_cols(7)
ws['G1'].value = "Time for a New Car?"

for col in range(1,10):
    ws[get_column_letter(col)+'1'].font = Font(bold= True, size=13)



for row in ws.iter_rows(min_row=2):
    miles_leftv = row[5].value
    max_miles = row[4].value
    new_car = row[6].row
    for i in ws.iter_rows():
        if int((row[5].value)) < (0):
            ws.cell(row=new_car,column=7).value = ("Yes")
            ws.cell(row=new_car,column=7).font = Font(color="00FF0000",bold=True )


        else:
            ws.cell(row=new_car,column=7).value = "No"


for letter in ['G','H']:
    max_width = 0
    for row_number in range(1,ws.max_row+1):
        if len(str(ws[f'{letter}{row_number}'].value)) > max_width:
            max_width = len(str(ws[f'{letter}{row_number}'].value))
    ws.column_dimensions[letter].width = max_width + 3



wb.save("Finished_Master_Report.xlsx")